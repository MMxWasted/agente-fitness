from copy import copy
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.schemas.body_measurement_import import BodyMeasurementImportPreview
from app.services.body_measurement_imports.workbook import (
    BodyMeasurementWorkbookAdapterV1,
    InvalidWorkbookError,
    UnsupportedWorkbookError,
    UploadTooLargeError,
    WorkbookPolicy,
    _parse_decimal,
    _validate_archive_entry,
    validate_xlsx_archive,
)

FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "body_measurements"
    / "body_measurements_format_v1.xlsx"
)
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
POLICY = WorkbookPolicy(
    max_file_size_bytes=5 * 1024 * 1024,
    max_zip_entries=512,
    max_uncompressed_size_bytes=25 * 1024 * 1024,
)


def fixture_bytes() -> bytes:
    return FIXTURE.read_bytes()


def workbook_bytes_with_change(
    cell_reference: str,
    value: object,
) -> bytes:
    workbook = load_workbook(FIXTURE, data_only=False, keep_links=False)
    worksheet = workbook["Revisiones"]
    worksheet[cell_reference] = value
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def preview(
    content: bytes | None = None,
) -> BodyMeasurementImportPreview:
    workbook_content = content or fixture_bytes()
    metadata = validate_xlsx_archive(
        workbook_content,
        filename="measurements.xlsx",
        content_type=XLSX_CONTENT_TYPE,
        policy=POLICY,
    )
    return BodyMeasurementWorkbookAdapterV1().preview(
        workbook_content,
        archive_metadata=metadata,
    )


def test_fixture_preserves_the_supported_v1_technical_structure() -> None:
    workbook = load_workbook(FIXTURE, data_only=False, keep_links=False)
    worksheet = workbook["Revisiones"]

    assert workbook.sheetnames == ["Revisiones"]
    assert worksheet.calculate_dimension() == "A1:F36"
    assert {str(item) for item in worksheet.merged_cells.ranges} == {
        "A1:F1",
        "A2:F2",
        "A5:A14",
        "A16:A23",
        "A25:A34",
    }
    assert [worksheet.cell(4, column).value for column in range(1, 4)] == [
        "Categoría",
        "Métrica",
        "Unidad",
    ]
    assert all(cell.data_type != "f" for row in worksheet.iter_rows() for cell in row)
    assert all(
        cell.comment is None and cell.hyperlink is None
        for row in worksheet.iter_rows()
        for cell in row
    )
    assert workbook.properties.creator == "openpyxl"
    assert workbook.properties.title is None
    assert workbook.properties.subject is None
    workbook.close()


def test_fixture_preview_detects_sections_revisions_and_metrics() -> None:
    result = preview()

    assert result.adapter_version == "body-measurements-v1"
    assert result.metadata.supported_sheet == "Revisiones"
    assert result.metadata.used_rows == 36
    assert result.metadata.used_columns == 6
    assert result.totals.revision_count == 3
    assert result.totals.recognized_metric_values == 83
    assert {
        metric.category for revision in result.revisions for metric in revision.metrics
    } == {"bioimpedance", "skinfold", "circumference"}
    assert len(result.unknown_metrics) == 1
    assert result.unknown_metrics[0].original_label == "Índice experimental"


def test_fixture_preview_handles_dates_decimals_units_and_laterality() -> None:
    result = preview()
    first_revision, ambiguous_revision, _ = result.revisions

    assert first_revision.normalized_date is not None
    assert first_revision.normalized_date.isoformat() == "2026-01-15"
    assert ambiguous_revision.raw_date == "06-03"
    assert ambiguous_revision.normalized_date is None
    assert ambiguous_revision.inferred_year == 2026
    assert any(
        issue.code == "revision_year_required" and issue.blocking
        for issue in result.errors
    )

    metrics = {
        (metric.code, metric.side): metric for metric in ambiguous_revision.metrics
    }
    assert metrics[("body_weight", "none")].value == "72.4"
    assert metrics[("body_weight", "none")].unit == "kg"
    assert metrics[("body_weight", "none")].unit_source == "excel"
    assert metrics[("quadriceps_skinfold", "left")].value == "13.8"
    assert metrics[("quadriceps_skinfold", "right")].value == "14"
    assert metrics[("visceral_fat_level", "none")].value == "0"
    assert ("bone_mass", "none") not in metrics
    assert any(
        cell.reference == "E11" and cell.reason == "empty_metric_value"
        for cell in result.ignored_cells
    )


def test_controlled_alias_is_recognized() -> None:
    result = preview(workbook_bytes_with_change("B5", "Peso"))

    assert any(metric.code == "body_weight" for metric in result.revisions[0].metrics)


def test_unit_mismatch_is_visible_and_blocking() -> None:
    result = preview(workbook_bytes_with_change("C5", "lb"))
    weight = next(
        metric for metric in result.revisions[0].metrics if metric.code == "body_weight"
    )

    assert weight.unit is None
    assert weight.unit_source == "unresolved"
    assert any(
        issue.code == "unit_mismatch"
        and issue.metric_code == "body_weight"
        and issue.blocking
        for issue in result.errors
    )


@pytest.mark.parametrize("value", ["31-02-2026", "2099-01-01"])
def test_impossible_and_future_revision_dates_are_rejected(value: str) -> None:
    with pytest.raises(InvalidWorkbookError):
        preview(workbook_bytes_with_change("D4", value))


@pytest.mark.parametrize(
    ("value", "expected_error"),
    [
        (True, "boolean_measurement_not_allowed"),
        ("NaN", "non_finite_measurement"),
        ("Infinity", "non_finite_measurement"),
        ("not-a-number", "invalid_number"),
    ],
)
def test_decimal_parser_rejects_unsafe_values(
    value: object,
    expected_error: str,
) -> None:
    parsed, _, error = _parse_decimal(value)

    assert parsed is None
    assert error == expected_error


def test_decimal_parser_warns_about_ambiguous_thousands_separator() -> None:
    parsed, warning, error = _parse_decimal("1,234")

    assert str(parsed) == "1.234"
    assert warning == "ambiguous_numeric_separator"
    assert error is None


def test_body_value_formula_is_rejected_without_execution() -> None:
    content = workbook_bytes_with_change("D5", "=1+1")

    with pytest.raises(
        InvalidWorkbookError,
        match="Formulas are not allowed",
    ):
        preview(content)


def test_protected_worksheet_is_rejected() -> None:
    workbook = load_workbook(FIXTURE, data_only=False, keep_links=False)
    workbook["Revisiones"].protection.sheet = True
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(InvalidWorkbookError, match="Protected worksheets"):
        preview(output.getvalue())


def test_fingerprint_is_stable_for_styles_and_changes_for_values() -> None:
    original = preview()
    workbook = load_workbook(FIXTURE, data_only=False, keep_links=False)
    font = copy(workbook["Revisiones"]["D5"].font)
    font.bold = True
    workbook["Revisiones"]["D5"].font = font
    restyled_output = BytesIO()
    workbook.save(restyled_output)
    workbook.close()

    restyled = preview(restyled_output.getvalue())
    changed = preview(workbook_bytes_with_change("D5", 73.1))

    assert restyled.fingerprint == original.fingerprint
    assert changed.fingerprint != original.fingerprint


def test_archive_validation_rejects_extension_mime_and_invalid_zip() -> None:
    content = fixture_bytes()

    with pytest.raises(UnsupportedWorkbookError, match=".xlsx"):
        validate_xlsx_archive(
            content,
            filename="measurements.xlsm",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )
    with pytest.raises(UnsupportedWorkbookError, match="content type"):
        validate_xlsx_archive(
            content,
            filename="measurements.xlsx",
            content_type="text/csv",
            policy=POLICY,
        )
    with pytest.raises(UnsupportedWorkbookError, match="valid XLSX"):
        validate_xlsx_archive(
            b"not a zip",
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )


def test_mime_is_auxiliary_when_the_xlsx_structure_is_valid() -> None:
    metadata = validate_xlsx_archive(
        fixture_bytes(),
        filename="measurements.xlsx",
        content_type="application/x-upload",
        policy=POLICY,
    )

    assert metadata.content_type_signal == "generic"


def test_archive_validation_rejects_macro_parts_and_unsafe_xml() -> None:
    macro_content = BytesIO(fixture_bytes())
    with ZipFile(macro_content, "a", ZIP_DEFLATED) as archive:
        archive.writestr("xl/vbaProject.bin", b"synthetic-macro-marker")
    with pytest.raises(UnsupportedWorkbookError, match="Macros"):
        validate_xlsx_archive(
            macro_content.getvalue(),
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )

    unsafe_xml_content = BytesIO(fixture_bytes())
    with ZipFile(unsafe_xml_content, "a", ZIP_DEFLATED) as archive:
        archive.writestr(
            "customXml/unsafe.xml",
            '<!DOCTYPE root [<!ENTITY x "unsafe">]><root>&x;</root>',
        )
    with pytest.raises(InvalidWorkbookError, match="Unsafe XML"):
        validate_xlsx_archive(
            unsafe_xml_content.getvalue(),
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )


def test_archive_validation_rejects_excessive_file_and_package_limits() -> None:
    content = fixture_bytes()
    with pytest.raises(UploadTooLargeError):
        validate_xlsx_archive(
            content,
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=WorkbookPolicy(
                max_file_size_bytes=100,
                max_zip_entries=512,
                max_uncompressed_size_bytes=25 * 1024 * 1024,
            ),
        )
    with pytest.raises(InvalidWorkbookError, match="too many entries"):
        validate_xlsx_archive(
            content,
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=WorkbookPolicy(
                max_file_size_bytes=5 * 1024 * 1024,
                max_zip_entries=2,
                max_uncompressed_size_bytes=25 * 1024 * 1024,
            ),
        )
    with pytest.raises(InvalidWorkbookError, match="uncompressed"):
        validate_xlsx_archive(
            content,
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=WorkbookPolicy(
                max_file_size_bytes=5 * 1024 * 1024,
                max_zip_entries=512,
                max_uncompressed_size_bytes=100,
            ),
        )


def test_archive_validation_rejects_missing_structure_and_unsafe_paths() -> None:
    incomplete = BytesIO()
    with ZipFile(incomplete, "w", ZIP_DEFLATED) as archive:
        archive.writestr("placeholder.txt", "not xlsx")
    with pytest.raises(InvalidWorkbookError, match="missing required"):
        validate_xlsx_archive(
            incomplete.getvalue(),
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )

    unsafe = BytesIO()
    with ZipFile(unsafe, "w", ZIP_DEFLATED) as archive:
        archive.writestr("../unsafe.xml", "<root />")
    with pytest.raises(InvalidWorkbookError, match="unsafe path"):
        validate_xlsx_archive(
            unsafe.getvalue(),
            filename="measurements.xlsx",
            content_type=XLSX_CONTENT_TYPE,
            policy=POLICY,
        )


def test_archive_validation_rejects_encrypted_entries() -> None:
    entry = ZipInfo("encrypted.xml")
    entry.flag_bits |= 0x1

    with pytest.raises(InvalidWorkbookError, match="Encrypted"):
        _validate_archive_entry(entry)
