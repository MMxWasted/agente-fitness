import hashlib
import json
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Literal

from openpyxl import DEFUSEDXML, load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import (  # type: ignore[import-untyped]
    Worksheet,
)

from app.schemas.body_measurement_import import (
    BodyMeasurementImportPreview,
    IgnoredCellPreview,
    MeasurementPreview,
    MeasurementUnit,
    PreviewIssue,
    PreviewTotals,
    RevisionPreview,
    TechnicalMetadata,
    UnknownMetricPreview,
)
from app.services.body_measurement_imports.catalog import (
    CATEGORY_ALIASES,
    METRICS_BY_ALIAS,
    UNIT_ALIASES,
    MetricDefinition,
    limited_label,
    normalize_label,
    split_laterality,
)

ADAPTER_VERSION: Literal["body-measurements-v1"] = "body-measurements-v1"
SUPPORTED_SHEET = "Revisiones"
_HEADER_LABELS = ("categoria", "metrica", "unidad")
_REQUIRED_ARCHIVE_ENTRIES = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
    }
)
_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_GENERIC_CONTENT_TYPES = {"application/octet-stream", "application/zip", ""}
_KNOWN_UNSUPPORTED_CONTENT_TYPES = {
    "application/vnd.ms-excel",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "text/csv",
}
_DATE_WITH_YEAR_FORMATS = (
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
)
_DATE_WITHOUT_YEAR = re.compile(r"^(?P<day>\d{1,2})[-/](?P<month>\d{1,2})$")
_WINDOWS_ABSOLUTE_PATH = re.compile(r"^[A-Za-z]:")


class BodyMeasurementImportError(Exception):
    """Base class for safe upload errors."""


class UploadTooLargeError(BodyMeasurementImportError):
    pass


class UnsupportedWorkbookError(BodyMeasurementImportError):
    pass


class InvalidWorkbookError(BodyMeasurementImportError):
    pass


@dataclass(frozen=True)
class WorkbookPolicy:
    max_file_size_bytes: int
    max_zip_entries: int
    max_uncompressed_size_bytes: int


@dataclass(frozen=True)
class ArchiveMetadata:
    entry_count: int
    uncompressed_size_bytes: int
    content_type_signal: Literal["xlsx", "generic", "missing"]


@dataclass(frozen=True)
class _RevisionColumn:
    column: int
    raw_date: str
    normalized_date: date | None
    label: str
    date_status: Literal["resolved", "missing_year"]
    inferred_year: int | None


def validate_xlsx_archive(
    content: bytes,
    *,
    filename: str | None,
    content_type: str | None,
    policy: WorkbookPolicy,
) -> ArchiveMetadata:
    if len(content) > policy.max_file_size_bytes:
        raise UploadTooLargeError("The uploaded file exceeds the configured limit")

    extension = Path(filename or "").suffix.casefold()
    if extension != ".xlsx":
        raise UnsupportedWorkbookError("Only macro-free .xlsx files are supported")

    normalized_content_type = (content_type or "").split(";", maxsplit=1)[0].strip()
    if normalized_content_type == _XLSX_CONTENT_TYPE:
        content_type_signal: Literal["xlsx", "generic", "missing"] = "xlsx"
    elif not normalized_content_type:
        content_type_signal = "missing"
    elif normalized_content_type in _KNOWN_UNSUPPORTED_CONTENT_TYPES:
        raise UnsupportedWorkbookError(
            "The uploaded content type is not compatible with .xlsx"
        )
    elif normalized_content_type in _GENERIC_CONTENT_TYPES:
        content_type_signal = "missing" if not normalized_content_type else "generic"
    else:
        content_type_signal = "generic"

    if not content.startswith(b"PK\x03\x04"):
        raise UnsupportedWorkbookError("The uploaded file is not a valid XLSX package")

    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > policy.max_zip_entries:
                raise InvalidWorkbookError("The XLSX package contains too many entries")

            names = [entry.filename for entry in entries]
            if len(names) != len(set(names)):
                raise InvalidWorkbookError(
                    "The XLSX package contains duplicate entries"
                )
            for entry in entries:
                _validate_archive_entry(entry)

            uncompressed_size = sum(entry.file_size for entry in entries)
            if uncompressed_size > policy.max_uncompressed_size_bytes:
                raise InvalidWorkbookError(
                    "The uncompressed XLSX package exceeds the configured limit"
                )

            available_entries = set(names)
            if not _REQUIRED_ARCHIVE_ENTRIES.issubset(available_entries):
                raise InvalidWorkbookError(
                    "The XLSX package is missing required workbook data"
                )
            if not any(
                name.startswith("xl/worksheets/") and name.endswith(".xml")
                for name in names
            ):
                raise InvalidWorkbookError(
                    "The XLSX package does not contain a worksheet"
                )
            if any(
                name.casefold().endswith("vbaproject.bin")
                or name.startswith("xl/externalLinks/")
                for name in names
            ):
                raise UnsupportedWorkbookError(
                    "Macros and external workbook links are not supported"
                )

            for entry in entries:
                if entry.filename.endswith((".xml", ".rels")):
                    xml_content = archive.read(entry)
                    upper_xml = xml_content.upper()
                    if b"<!DOCTYPE" in upper_xml or b"<!ENTITY" in upper_xml:
                        raise InvalidWorkbookError(
                            "Unsafe XML declarations are not supported"
                        )
    except zipfile.BadZipFile as error:
        raise UnsupportedWorkbookError(
            "The uploaded file is not a valid XLSX package"
        ) from error

    return ArchiveMetadata(
        entry_count=len(entries),
        uncompressed_size_bytes=uncompressed_size,
        content_type_signal=content_type_signal,
    )


def _validate_archive_entry(entry: zipfile.ZipInfo) -> None:
    name = entry.filename
    path = PurePosixPath(name)
    if (
        not name
        or name.startswith(("/", "\\"))
        or "\\" in name
        or _WINDOWS_ABSOLUTE_PATH.match(name)
        or ".." in path.parts
    ):
        raise InvalidWorkbookError("The XLSX package contains an unsafe path")
    if entry.flag_bits & 0x1:
        raise InvalidWorkbookError("Encrypted XLSX packages are not supported")


class BodyMeasurementWorkbookAdapterV1:
    version = ADAPTER_VERSION

    def __init__(self, *, today: date | None = None) -> None:
        self._today = today or datetime.now(UTC).date()

    def preview(
        self,
        content: bytes,
        *,
        archive_metadata: ArchiveMetadata,
    ) -> BodyMeasurementImportPreview:
        if not DEFUSEDXML:
            raise RuntimeError("The safe XML parser is unavailable")

        try:
            workbook = load_workbook(
                BytesIO(content),
                read_only=False,
                data_only=False,
                keep_vba=False,
                keep_links=False,
            )
        except Exception as error:
            raise InvalidWorkbookError(
                "The XLSX workbook could not be read safely"
            ) from error

        try:
            self._validate_workbook(workbook)
            worksheet = workbook[SUPPORTED_SHEET]
            return self._build_preview(
                worksheet,
                file_size_bytes=len(content),
                sheet_count=len(workbook.sheetnames),
                archive_metadata=archive_metadata,
            )
        finally:
            workbook.close()

    def _validate_workbook(self, workbook: object) -> None:
        sheetnames = getattr(workbook, "sheetnames", [])
        if sheetnames != [SUPPORTED_SHEET]:
            raise InvalidWorkbookError(
                "The workbook does not match the supported format"
            )
        security = getattr(workbook, "security", None)
        if security is not None and any(
            bool(getattr(security, attribute, False))
            for attribute in ("lockStructure", "lockWindows", "lockRevision")
        ):
            raise InvalidWorkbookError("Protected workbooks are not supported")

        worksheet = workbook[SUPPORTED_SHEET]  # type: ignore[index]
        if worksheet.protection.sheet:
            raise InvalidWorkbookError("Protected worksheets are not supported")
        if worksheet.max_row > 500 or worksheet.max_column > 100:
            raise InvalidWorkbookError(
                "The workbook dimensions exceed the supported format"
            )

    def _build_preview(
        self,
        worksheet: Worksheet,
        *,
        file_size_bytes: int,
        sheet_count: int,
        archive_metadata: ArchiveMetadata,
    ) -> BodyMeasurementImportPreview:
        header_row = self._find_header_row(worksheet)
        revision_columns, warnings, errors = self._read_revision_columns(
            worksheet,
            header_row,
        )
        metrics_by_revision: dict[int, list[MeasurementPreview]] = {
            revision.column: [] for revision in revision_columns
        }
        unknown_metrics: list[UnknownMetricPreview] = []
        ignored_cells: list[IgnoredCellPreview] = []
        current_category_label = ""

        for row in range(header_row + 1, worksheet.max_row + 1):
            category_cell = worksheet.cell(row, 1)
            metric_cell = worksheet.cell(row, 2)
            unit_cell = worksheet.cell(row, 3)
            if category_cell.value is not None:
                current_category_label = limited_label(category_cell.value)

            if metric_cell.value is None:
                if any(
                    worksheet.cell(row, revision.column).value is not None
                    for revision in revision_columns
                ):
                    raise InvalidWorkbookError(
                        "A measurement row is missing its metric label"
                    )
                ignored_cells.append(
                    IgnoredCellPreview(
                        reference=f"A{row}",
                        reason="separator_row",
                    )
                )
                continue
            if metric_cell.data_type == "f" or unit_cell.data_type == "f":
                raise InvalidWorkbookError(
                    "Formulas are not allowed in measurement labels or units"
                )
            if any(
                worksheet.cell(row, revision.column).data_type == "f"
                for revision in revision_columns
            ):
                raise InvalidWorkbookError(
                    "Formulas are not allowed as body measurement values"
                )

            original_label = limited_label(metric_cell.value)
            normalized_metric, side = split_laterality(original_label)
            definition = METRICS_BY_ALIAS.get(normalized_metric)
            normalized_category = CATEGORY_ALIASES.get(
                normalize_label(current_category_label)
            )

            if (
                definition is None
                or normalized_category is None
                or definition.category != normalized_category
                or side not in definition.allowed_sides
            ):
                populated_count = sum(
                    worksheet.cell(row, revision.column).value is not None
                    for revision in revision_columns
                )
                unknown_metrics.append(
                    UnknownMetricPreview(
                        category_label=limited_label(
                            current_category_label,
                            maximum=40,
                        ),
                        side=side,
                        original_label=original_label,
                        populated_revision_count=populated_count,
                    )
                )
                if normalized_category is None:
                    ignored_cells.append(
                        IgnoredCellPreview(
                            reference=category_cell.coordinate,
                            reason="unsupported_section",
                        )
                    )
                continue

            unit, unit_source, unit_issue = self._read_unit(
                unit_cell,
                definition,
            )
            if unit_issue is not None:
                errors.append(unit_issue)

            for revision in revision_columns:
                value_cell = worksheet.cell(row, revision.column)
                if value_cell.value is None or (
                    isinstance(value_cell.value, str) and not value_cell.value.strip()
                ):
                    ignored_cells.append(
                        IgnoredCellPreview(
                            reference=value_cell.coordinate,
                            reason="empty_metric_value",
                        )
                    )
                    continue

                parsed_value, numeric_warning, numeric_error = _parse_decimal(
                    value_cell.value
                )
                if numeric_warning is not None:
                    warnings.append(
                        PreviewIssue(
                            code=numeric_warning,
                            message=(
                                "El separador numérico es ambiguo y deberá "
                                "revisarse antes de confirmar."
                            ),
                            blocking=False,
                            revision_label=revision.label,
                            metric_code=definition.code,
                        )
                    )
                if numeric_error is not None or parsed_value is None:
                    errors.append(
                        PreviewIssue(
                            code=numeric_error or "invalid_number",
                            message=(
                                "El valor de una medición no es un número "
                                "decimal finito válido."
                            ),
                            blocking=True,
                            revision_label=revision.label,
                            metric_code=definition.code,
                        )
                    )
                    continue
                if not definition.minimum <= parsed_value <= definition.maximum:
                    errors.append(
                        PreviewIssue(
                            code="measurement_out_of_range",
                            message=(
                                "El valor comunicado queda fuera del rango "
                                "básico admitido para esta métrica."
                            ),
                            blocking=True,
                            revision_label=revision.label,
                            metric_code=definition.code,
                        )
                    )
                    continue

                metrics_by_revision[revision.column].append(
                    MeasurementPreview(
                        code=definition.code,
                        category=definition.category,
                        side=side,
                        value=_canonical_decimal(parsed_value),
                        unit=unit,
                        unit_source=unit_source,
                        original_label=original_label,
                    )
                )

        revisions = [
            RevisionPreview(
                raw_date=revision.raw_date,
                normalized_date=revision.normalized_date,
                label=revision.label,
                date_status=revision.date_status,
                inferred_year=revision.inferred_year,
                metrics=sorted(
                    metrics_by_revision[revision.column],
                    key=lambda metric: (
                        metric.category,
                        metric.code,
                        metric.side,
                    ),
                ),
            )
            for revision in revision_columns
        ]
        fingerprint = self._fingerprint(
            revisions,
            unknown_metrics,
            warnings,
            errors,
        )
        totals = PreviewTotals(
            revision_count=len(revisions),
            recognized_metric_values=sum(
                len(revision.metrics) for revision in revisions
            ),
            unknown_metric_rows=len(unknown_metrics),
            ignored_cells=len(ignored_cells),
            warning_count=len(warnings),
            error_count=len(errors),
            has_blocking_errors=any(issue.blocking for issue in errors),
        )
        return BodyMeasurementImportPreview(
            adapter_version=ADAPTER_VERSION,
            fingerprint=f"sha256:{fingerprint}",
            metadata=TechnicalMetadata(
                file_size_bytes=file_size_bytes,
                sheet_count=sheet_count,
                supported_sheet=SUPPORTED_SHEET,
                used_rows=worksheet.max_row,
                used_columns=worksheet.max_column,
                zip_entry_count=archive_metadata.entry_count,
                uncompressed_size_bytes=archive_metadata.uncompressed_size_bytes,
                content_type_signal=archive_metadata.content_type_signal,
            ),
            revisions=revisions,
            warnings=warnings,
            errors=errors,
            unknown_metrics=unknown_metrics,
            ignored_cells=ignored_cells,
            totals=totals,
        )

    def _find_header_row(self, worksheet: Worksheet) -> int:
        for row in range(1, min(worksheet.max_row, 20) + 1):
            labels = tuple(
                normalize_label(str(worksheet.cell(row, column).value or ""))
                for column in range(1, 4)
            )
            if labels == _HEADER_LABELS:
                if worksheet.max_column < 4:
                    break
                return row
        raise InvalidWorkbookError(
            "The workbook does not contain the supported measurement headers"
        )

    def _read_revision_columns(
        self,
        worksheet: Worksheet,
        header_row: int,
    ) -> tuple[list[_RevisionColumn], list[PreviewIssue], list[PreviewIssue]]:
        full_years: set[int] = set()
        parsed_headers: list[tuple[int, object, date | None]] = []
        for column in range(4, worksheet.max_column + 1):
            cell = worksheet.cell(header_row, column)
            if cell.data_type == "f":
                raise InvalidWorkbookError(
                    "Formulas are not allowed in revision headers"
                )
            parsed_date = self._parse_full_date(cell.value)
            if parsed_date is not None:
                full_years.add(parsed_date.year)
            parsed_headers.append((column, cell.value, parsed_date))

        inferred_year = next(iter(full_years)) if len(full_years) == 1 else None
        revisions: list[_RevisionColumn] = []
        warnings: list[PreviewIssue] = []
        errors: list[PreviewIssue] = []
        for column, value, parsed_date in parsed_headers:
            raw_date = _safe_raw_date(value)
            if parsed_date is not None:
                revisions.append(
                    _RevisionColumn(
                        column=column,
                        raw_date=raw_date,
                        normalized_date=parsed_date,
                        label=parsed_date.isoformat(),
                        date_status="resolved",
                        inferred_year=None,
                    )
                )
                continue

            match = _DATE_WITHOUT_YEAR.fullmatch(raw_date)
            if match is None:
                raise InvalidWorkbookError(
                    "A revision column has an invalid or missing date"
                )
            day = int(match.group("day"))
            month = int(match.group("month"))
            try:
                date(2000, month, day)
            except ValueError as error:
                raise InvalidWorkbookError(
                    "A revision column has an impossible date"
                ) from error

            label = f"{day:02d}-{month:02d} (año pendiente)"
            errors.append(
                PreviewIssue(
                    code="revision_year_required",
                    message=(
                        "La fecha de revisión no incluye año y deberá "
                        "resolverse antes de confirmar."
                    ),
                    blocking=True,
                    revision_label=label,
                )
            )
            if inferred_year is not None:
                warnings.append(
                    PreviewIssue(
                        code="revision_year_inferred_candidate",
                        message=(
                            "El resto del libro comparte un año; se muestra "
                            "solo como propuesta, no como fecha resuelta."
                        ),
                        blocking=False,
                        revision_label=label,
                    )
                )
            revisions.append(
                _RevisionColumn(
                    column=column,
                    raw_date=raw_date,
                    normalized_date=None,
                    label=label,
                    date_status="missing_year",
                    inferred_year=inferred_year,
                )
            )

        if not revisions:
            raise InvalidWorkbookError("The workbook does not contain revision columns")
        resolved_dates = [
            revision.normalized_date
            for revision in revisions
            if revision.normalized_date is not None
        ]
        if len(resolved_dates) != len(set(resolved_dates)):
            errors.append(
                PreviewIssue(
                    code="duplicate_revision_date",
                    message="Dos columnas representan la misma fecha de revisión.",
                    blocking=True,
                )
            )
        return revisions, warnings, errors

    def _parse_full_date(self, value: object) -> date | None:
        parsed: date | None = None
        if isinstance(value, datetime):
            parsed = value.date()
        elif isinstance(value, date):
            parsed = value
        elif isinstance(value, str):
            raw_value = value.strip()
            for date_format in _DATE_WITH_YEAR_FORMATS:
                try:
                    parsed = datetime.strptime(raw_value, date_format).date()
                    break
                except ValueError:
                    continue
        if parsed is not None and parsed > self._today:
            raise InvalidWorkbookError("Future revision dates are not supported")
        return parsed

    def _read_unit(
        self,
        cell: Cell,
        definition: MetricDefinition,
    ) -> tuple[
        MeasurementUnit | None,
        Literal["excel", "adapter_v1", "unresolved"],
        PreviewIssue | None,
    ]:
        if cell.value is None or not str(cell.value).strip():
            return definition.unit, "adapter_v1", None
        normalized_unit = UNIT_ALIASES.get(normalize_label(str(cell.value)))
        if normalized_unit == definition.unit:
            return definition.unit, "excel", None
        return (
            None,
            "unresolved",
            PreviewIssue(
                code="unit_mismatch",
                message=(
                    "La unidad declarada no coincide con la definición "
                    "versionada de la métrica."
                ),
                blocking=True,
                metric_code=definition.code,
            ),
        )

    def _fingerprint(
        self,
        revisions: list[RevisionPreview],
        unknown_metrics: list[UnknownMetricPreview],
        warnings: list[PreviewIssue],
        errors: list[PreviewIssue],
    ) -> str:
        payload = {
            "adapter_version": ADAPTER_VERSION,
            "revisions": sorted(
                (
                    {
                        "raw_date": revision.raw_date,
                        "normalized_date": (
                            revision.normalized_date.isoformat()
                            if revision.normalized_date
                            else None
                        ),
                        "date_status": revision.date_status,
                        "inferred_year": revision.inferred_year,
                        "metrics": sorted(
                            (
                                {
                                    "code": metric.code,
                                    "side": metric.side,
                                    "value": metric.value,
                                    "unit": metric.unit,
                                    "unit_source": metric.unit_source,
                                }
                                for metric in revision.metrics
                            ),
                            key=lambda metric: (
                                metric["code"],
                                metric["side"],
                                metric["value"],
                            ),
                        ),
                    }
                    for revision in revisions
                ),
                key=lambda revision: (
                    revision["normalized_date"] or "",
                    revision["raw_date"],
                ),
            ),
            "unknown_metrics": sorted(
                (
                    {
                        "category": normalize_label(metric.category_label),
                        "label": normalize_label(metric.original_label),
                        "side": metric.side,
                        "populated": metric.populated_revision_count,
                    }
                    for metric in unknown_metrics
                ),
                key=lambda metric: (
                    metric["category"],
                    metric["label"],
                    metric["side"],
                ),
            ),
            "ambiguities": sorted(
                (
                    {
                        "code": issue.code,
                        "revision": issue.revision_label,
                        "metric": issue.metric_code,
                        "blocking": issue.blocking,
                    }
                    for issue in (*warnings, *errors)
                ),
                key=lambda issue: (
                    issue["code"],
                    issue["revision"] or "",
                    issue["metric"] or "",
                ),
            ),
        }
        canonical_json = json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        )
        return hashlib.sha256(canonical_json.encode("utf-8")).hexdigest()


def _parse_decimal(
    value: object,
) -> tuple[Decimal | None, str | None, str | None]:
    if isinstance(value, bool):
        return None, None, "boolean_measurement_not_allowed"
    if isinstance(value, float) and not math.isfinite(value):
        return None, None, "non_finite_measurement"

    ambiguous_separator = False
    if isinstance(value, str):
        normalized_value = value.strip().replace("\u00a0", " ")
        if not normalized_value:
            return None, None, "invalid_number"
        if " " in normalized_value:
            return None, None, "ambiguous_numeric_separator"
        if "," in normalized_value and "." in normalized_value:
            return None, None, "ambiguous_numeric_separator"
        separator_match = re.fullmatch(r"[+-]?\d{1,3}([,.])\d{3}", normalized_value)
        ambiguous_separator = separator_match is not None
        normalized_value = normalized_value.replace(",", ".")
    elif isinstance(value, (int, float, Decimal)):
        normalized_value = str(value)
    else:
        return None, None, "invalid_number"

    try:
        decimal_value = Decimal(normalized_value)
    except InvalidOperation:
        return None, None, "invalid_number"
    if not decimal_value.is_finite():
        return None, None, "non_finite_measurement"
    warning = "ambiguous_numeric_separator" if ambiguous_separator else None
    return decimal_value, warning, None


def _canonical_decimal(value: Decimal) -> str:
    if value == 0:
        return "0"
    canonical = format(value.normalize(), "f")
    if "." in canonical:
        canonical = canonical.rstrip("0").rstrip(".")
    return canonical


def _safe_raw_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return limited_label(value, maximum=30)
    return ""
